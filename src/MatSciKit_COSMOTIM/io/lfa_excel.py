"""
LFA Excel data reader.

Reads laser flash analysis diffusivity data from Excel files
(e.g., UT LFA instrument export format).
"""

from __future__ import annotations

from pathlib import Path

import numpy as np


def read(filepath: str, sheet_name: str | int = 0, skip_rows: int = 1) -> np.ndarray:
    """Read LFA diffusivity data from an Excel file.

    Expected format after skipping header:
    columns [Temperature (°C), T_error, Diffusivity (mm²/s), Diff_error].

    Temperature is converted from °C to K.

    Parameters
    ----------
    filepath : str
        Path to the Excel file (.xlsx).
    sheet_name : str or int, optional
        Sheet name or index (default 0).
    skip_rows : int, optional
        Number of header rows to skip (default 1).

    Returns
    -------
    data : np.ndarray
        Array with columns [Temperature (K), Diffusivity (mm²/s), Diffusivity_error].

    Raises
    ------
    FileNotFoundError
        If the file does not exist.
    ImportError
        If openpyxl is not installed.
    """
    filepath = Path(filepath)
    if not filepath.exists():
        raise FileNotFoundError(f"File not found: {filepath}")

    try:
        import openpyxl
    except ImportError as exc:
        raise ImportError(
            "openpyxl is required to read Excel files. Install with: pip install openpyxl"
        ) from exc

    wb = openpyxl.load_workbook(filepath, read_only=True, data_only=True)

    ws = wb.worksheets[sheet_name] if isinstance(sheet_name, int) else wb[sheet_name]

    rows = []
    for i, row in enumerate(ws.iter_rows(values_only=True)):
        if i < skip_rows:
            continue
        # Skip empty rows
        if row[0] is None:
            continue
        try:
            vals = [float(v) if v is not None else np.nan for v in row[:4]]
            rows.append(vals)
        except (ValueError, TypeError):
            continue

    wb.close()

    if not rows:
        raise ValueError(f"No numeric data found in sheet '{sheet_name}'")

    raw = np.array(rows)

    # Extract: Temperature (°C → K), Diffusivity, Diffusivity_error
    temp_k = raw[:, 0] + 273.15
    diffusivity = raw[:, 2]
    diff_error = raw[:, 3]

    result = np.column_stack([temp_k, diffusivity, diff_error])

    # Drop NaN rows
    mask = ~np.any(np.isnan(result), axis=1)
    result = result[mask]

    return result


def list_sheets(filepath: str) -> list[str]:
    """List available sheet names in an LFA Excel file.

    Parameters
    ----------
    filepath : str
        Path to the Excel file.

    Returns
    -------
    sheets : list of str
        Sheet names.
    """
    try:
        import openpyxl
    except ImportError as exc:
        raise ImportError(
            "openpyxl is required to read Excel files. Install with: pip install openpyxl"
        ) from exc

    wb = openpyxl.load_workbook(filepath, read_only=True)
    names = wb.sheetnames
    wb.close()
    return names
